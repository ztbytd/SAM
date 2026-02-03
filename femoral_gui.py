#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
股骨头坏死灰度分析 GUI应用
========================
功能：
- 上传图片 / 从DICOM校准 / 手动输入px/mm
- 自动/手动识别股骨头
- 画股骨头圆、r-6mm内圆
- 自动检测硬化带交点
- 计算灰度比值
- 保存结果图片

打包成exe: pyinstaller --onefile --windowed femoral_gui.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk, ImageDraw, ImageFont
import cv2
import numpy as np
import os
from datetime import datetime
from pathlib import Path
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ImageData:
    """单张图片数据类"""
    def __init__(self, path):
        self.path = path
        self.name = os.path.basename(path)
        self.analyzer = None
        self.results = []
        self.is_analyzed = False
        self.is_saved = False
        self.thumbnail = None
        self.px_per_mm_source = "(默认)"  # 校准来源
        # 保存分析数据的副本（避免引用问题）
        self.center = None
        self.radius = None
        self.inner_radius = None
        self.px_per_mm = None

    def to_dict(self):
        """转换为字典用于导出"""
        if not self.results:
            return {
                'filename': self.name,
                'status': '未分析',
                'ratio_1': '--',
                'ratio_2': '--',
                'avg_ratio': '--'
            }

        ratios = [r['ratio'] for r in self.results[:2]]
        while len(ratios) < 2:
            ratios.append(None)

        avg_ratio = np.mean([r['ratio'] for r in self.results]) if self.results else 0

        return {
            'filename': self.name,
            'status': '已分析' if self.is_analyzed else '未分析',
            'ratio_1': f"{ratios[0]:.4f}" if ratios[0] is not None else '--',
            'ratio_2': f"{ratios[1]:.4f}" if ratios[1] is not None else '--',
            'avg_ratio': f"{avg_ratio:.4f}" if avg_ratio > 0 else '--'
        }


class ImageProject:
    """图片项目管理类"""
    def __init__(self):
        self.images = []
        self.current_index = -1
        self.output_dir = None

    def add_images(self, paths):
        """添加多张图片"""
        for path in paths:
            if os.path.isfile(path):
                self.images.append(ImageData(path))

    def get_current_image(self):
        """获取当前图片"""
        if 0 <= self.current_index < len(self.images):
            return self.images[self.current_index]
        return None

    def set_current_index(self, index):
        """设置当前图片索引"""
        if 0 <= index < len(self.images):
            self.current_index = index
            return True
        return False

    def next_image(self):
        """下一张图片"""
        if self.current_index < len(self.images) - 1:
            self.current_index += 1
            return True
        return False

    def prev_image(self):
        """上一张图片"""
        if self.current_index > 0:
            self.current_index -= 1
            return True
        return False

    def get_analyzed_count(self):
        """获取已分析的图片数"""
        return sum(1 for img in self.images if img.is_analyzed)

    def get_saved_count(self):
        """获取已保存的图片数"""
        return sum(1 for img in self.images if img.is_saved)

    def create_output_dir(self, base_path=None):
        """创建输出文件夹"""
        if base_path is None:
            base_path = os.path.expanduser("~")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = os.path.join(base_path, f"femoral_analysis_{timestamp}")

        try:
            os.makedirs(self.output_dir, exist_ok=True)
            return self.output_dir
        except Exception as e:
            print(f"创建输出文件夹失败: {e}")
            return None

    def export_to_excel(self, filepath):
        """导出分析结果到Excel（只导出已分析的数据）"""
        if not EXCEL_AVAILABLE:
            return False, "需要安装openpyxl库: pip install openpyxl"

        # 只筛选已分析的图片
        analyzed_images = [img for img in self.images if img.is_analyzed]
        
        if not analyzed_images:
            return False, "没有已分析的数据可导出"

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "分析结果"

            # 设置列宽
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12

            # 标题行（只有2个交点）
            headers = ['文件名', '状态', '交点1比值', '交点2比值', '平均比值']
            ws.append(headers)

            # 设置标题样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 数据行（只导出已分析的）
            for img_data in analyzed_images:
                data_dict = img_data.to_dict()
                ws.append([
                    data_dict['filename'],
                    data_dict['status'],
                    data_dict['ratio_1'],
                    data_dict['ratio_2'],
                    data_dict['avg_ratio']
                ])

            # 设置数据行样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=2, max_row=len(analyzed_images)+1, min_col=1, max_col=5):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # 冻结标题行
            ws.freeze_panes = "A2"

            wb.save(filepath)
            return True, f"已导出 {len(analyzed_images)} 条分析数据到: {filepath}"
        except Exception as e:
            return False, f"导出失败: {e}"


class FemoralHeadAnalyzer:
    """股骨头分析核心类"""

    def __init__(self, px_per_mm=5.0):
        self.px_per_mm = px_per_mm
        self.image = None
        self.gray = None
        self.center = None
        self.radius = None
        self.inner_radius = None
        self.results = []
        
    def load_image(self, path):
        """加载图像（支持中文路径）"""
        # 使用 np.fromfile + cv2.imdecode 解决中文路径问题
        self.image = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
        if self.image is None:
            raise ValueError(f"无法加载图像: {path}")
        self.gray = cv2.cvtColor(self.image, cv2.COLOR_BGR2GRAY)
        return self.image
    
    def detect_femoral_head(self, search_roi=None, min_r=30, max_r=150):
        """自动检测股骨头"""
        if self.gray is None:
            return None
            
        if search_roi:
            x, y, w, h = search_roi
            search = self.gray[y:y+h, x:x+w]
            offset = (x, y)
        else:
            search = self.gray
            offset = (0, 0)
        
        blurred = cv2.GaussianBlur(search, (9, 9), 2)
        
        for p2 in [25, 30, 20, 35, 40]:
            circles = cv2.HoughCircles(
                blurred, cv2.HOUGH_GRADIENT, 1.2, 50,
                param1=50, param2=p2, minRadius=min_r, maxRadius=max_r
            )
            if circles is not None:
                c = max(circles[0], key=lambda x: x[2])
                self.center = (int(c[0]) + offset[0], int(c[1]) + offset[1])
                self.radius = int(c[2])
                return self.center, self.radius
        return None
    
    def set_femoral_head(self, center, radius):
        """手动设置股骨头"""
        self.center = center
        self.radius = radius
    
    def find_sclerotic_intersections(self, inner_radius, n_samples=72):
        """找硬化带交点"""
        if self.gray is None or self.center is None:
            return []
        
        cx, cy = self.center
        samples = []
        
        for i in range(n_samples):
            angle = 2 * np.pi * i / n_samples
            x = int(cx + inner_radius * np.cos(angle))
            y = int(cy + inner_radius * np.sin(angle))
            
            if 0 <= x < self.gray.shape[1] and 0 <= y < self.gray.shape[0]:
                w = 2
                region = self.gray[max(0,y-w):y+w+1, max(0,x-w):x+w+1]
                samples.append({
                    'angle': angle, 'deg': i * 360 / n_samples,
                    'x': x, 'y': y, 'bright': float(np.mean(region))
                })
        
        if not samples:
            return []
        
        brights = [s['bright'] for s in samples]
        mean_b, std_b = np.mean(brights), np.std(brights)
        
        # 找峰值
        peaks = []
        for i in range(len(samples)):
            prev_i = (i - 1) % len(samples)
            next_i = (i + 1) % len(samples)
            if (samples[i]['bright'] > samples[prev_i]['bright'] and 
                samples[i]['bright'] > samples[next_i]['bright'] and
                samples[i]['bright'] > mean_b + 0.3 * std_b):
                peaks.append(samples[i])
        
        peaks.sort(key=lambda x: x['bright'], reverse=True)
        
        # 确保峰值间隔足够
        final_peaks = []
        for p in peaks:
            is_far = True
            for fp in final_peaks:
                angle_diff = abs(p['deg'] - fp['deg'])
                if angle_diff > 180:
                    angle_diff = 360 - angle_diff
                if angle_diff < 40:
                    is_far = False
                    break
            if is_far:
                final_peaks.append(p)
            if len(final_peaks) >= 2:
                break
        
        return final_peaks
    
    def analyze(self, inner_offset_mm=6.0, roi_offset_mm=2.0, roi_radius_mm=1.0):
        """执行分析"""
        if self.gray is None or self.center is None or self.radius is None:
            return []

        cx, cy = self.center
        inner_offset_px = int(inner_offset_mm * self.px_per_mm)
        self.inner_radius = max(10, self.radius - inner_offset_px)
        roi_radius_px = max(3, int(roi_radius_mm * self.px_per_mm))
        # 偏移量 = 橙色点半径(4) + ROI半径，让测量圆边缘刚好挨着橙色点边缘
        orange_radius = 4
        offset_px = orange_radius + roi_radius_px

        intersections = self.find_sclerotic_intersections(self.inner_radius)

        self.results = []
        for inter in intersections:
            dx = inter['x'] - cx
            dy = inter['y'] - cy
            dist = np.sqrt(dx*dx + dy*dy)
            if dist > 0:
                ux, uy = dx/dist, dy/dist
            else:
                ux, uy = 0, -1

            # 硬化带ROI（向内偏移，靠近圆心）
            sclerotic_x = int(inter['x'] - ux * offset_px)
            sclerotic_y = int(inter['y'] - uy * offset_px)

            # 坏死区ROI（向外偏移，远离圆心，与硬化带对称）
            necrosis_x = int(inter['x'] + ux * offset_px)
            necrosis_y = int(inter['y'] + uy * offset_px)

            g_necrosis = self._calc_roi_gray(necrosis_x, necrosis_y, roi_radius_px)
            g_sclerotic = self._calc_roi_gray(sclerotic_x, sclerotic_y, roi_radius_px)

            ratio = g_sclerotic / g_necrosis if g_necrosis > 0 else 0

            self.results.append({
                'intersection': inter,
                'necrosis': (necrosis_x, necrosis_y),
                'sclerotic': (sclerotic_x, sclerotic_y),
                'roi_radius': roi_radius_px,
                'g_necrosis': g_necrosis,
                'g_sclerotic': g_sclerotic,
                'ratio': ratio
            })

        return self.results
    
    def _calc_roi_gray(self, cx, cy, r):
        """计算ROI平均灰度"""
        mask = np.zeros(self.gray.shape, np.uint8)
        cv2.circle(mask, (cx, cy), r, 255, -1)
        pixels = self.gray[mask > 0]
        return float(np.mean(pixels)) if len(pixels) > 0 else 0
    
    def draw_result(self):
        """绘制结果图像"""
        if self.image is None or self.center is None:
            return None
        
        vis = self.image.copy()
        cx, cy = self.center
        
        # 股骨头（绿色）
        cv2.circle(vis, self.center, self.radius, (0, 255, 0), 2)
        
        # 圆心（红色十字）
        cv2.line(vis, (cx-10, cy), (cx+10, cy), (0, 0, 255), 2)
        cv2.line(vis, (cx, cy-10), (cx, cy+10), (0, 0, 255), 2)
        
        # r-6mm内圆（青色）
        if self.inner_radius:
            cv2.circle(vis, self.center, self.inner_radius, (255, 255, 0), 2)
        
        # 交点和ROI
        for r in self.results:
            inter = r['intersection']
            # 交点（橙色）
            cv2.circle(vis, (inter['x'], inter['y']), 4, (0, 165, 255), -1)

            # 硬化带ROI（绿色，线宽1）
            cv2.circle(vis, r['sclerotic'], r['roi_radius'], (0, 255, 0), 1)

            # 坏死区ROI（红色，线宽1）
            cv2.circle(vis, r['necrosis'], r['roi_radius'], (0, 0, 255), 1)

        return vis


class FemoralHeadApp:
    """GUI应用主类"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("股骨头坏死灰度分析 - 多图片批处理")
        self.root.state('zoomed')  # 默认全屏
        self.root.configure(bg='#2b2b2b')

        # 配置滚动条样式 - 现代化设计
        style = ttk.Style()
        style.configure('Custom.Vertical.TScrollbar',
                       background='#2b2b2b',
                       troughcolor='#2b2b2b',
                       bordercolor='#2b2b2b',
                       darkcolor='#2b2b2b',
                       lightcolor='#2b2b2b',
                       arrowcolor='#888888',
                       relief='flat')

        # 项目管理
        self.project = ImageProject()

        # 分析器
        self.analyzer = FemoralHeadAnalyzer()

        # 变量
        self.image_path = None
        self.current_image = None
        self.display_image = None
        self.click_points = []  # 用于手动选择
        self.manual_mode = False  # 手动圆心模式
        self.intersection_mode = False  # 手动交点模式
        self.intersection_points = []  # 手动交点坐标
        self.necrosis_direction_point = None  # 坏死区方向点
        self.ruler_mode = False  # 标尺校准模式
        self.ruler_points = []   # 标尺校准点
        self.ruler_backup_image = None  # 标尺校准时的备份图像
        self.zoom_scale = 1.0  # 图片缩放比例
        self.close_btn_coords = None  # 关闭按钮坐标

        # 拖动相关变量
        self.drag_start_x = 0
        self.drag_start_y = 0
        self.image_offset_x = 0  # 图片偏移量
        self.image_offset_y = 0
        self.is_dragging = False
        self.click_on_image = False  # 是否点击在图片上

        # 缩放优化相关
        self._pil_cache = None       # 缓存颜色转换后的PIL图像
        self._cache_id = None        # 缓存标识（检测图像变化）
        self._zoom_timer = None      # 缩放防抖定时器

        # 参数变量
        self.px_per_mm = tk.DoubleVar(value=5.0)
        self.inner_offset = tk.DoubleVar(value=6.0)
        self.roi_radius = tk.DoubleVar(value=1.0)

        # 图片列表相关
        self.image_list_frame = None
        self.image_list_canvas = None
        self.image_list_scrollbar = None
        self.image_buttons = []
        self.selected_button = None

        self.setup_ui()

    def setup_image_list_panel(self, parent):
        """设置图片列表面板"""
        # 左侧面板框架 - 增加宽度到200
        list_panel = ttk.Frame(parent, width=200)
        list_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        list_panel.pack_propagate(False)

        # 标题
        ttk.Label(list_panel, text="图片列表", font=('Microsoft YaHei', 11, 'bold')).pack(pady=5)

        # 创建容器框架用于Canvas和滚动条
        self.list_container = tk.Frame(list_panel, bg='#2b2b2b')
        self.list_container.pack(fill=tk.BOTH, expand=True)

        # 创建带自定义滚动条的画布
        self.image_list_canvas = tk.Canvas(self.list_container, bg='#3b3b3b', highlightthickness=0, width=190)
        self.image_list_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 创建自定义滚动条 - 初始隐藏
        self.image_list_scrollbar_frame = tk.Frame(self.list_container, bg='#3b3b3b', width=8)
        # 不立即pack，等需要时再显示
        self.image_list_scrollbar_frame.pack_propagate(False)

        # 自定义滚动条轨道 - 圆角效果用更深的背景
        self.scrollbar_track = tk.Canvas(self.image_list_scrollbar_frame, bg='#3b3b3b',
                                         highlightthickness=0, width=6, cursor='arrow')
        self.scrollbar_track.pack(fill=tk.BOTH, expand=True, padx=1, pady=5)

        # 自定义滚动条滑块
        self.scrollbar_thumb = None
        self.scrollbar_thumb_id = None
        
        # 滚动条自动隐藏相关
        self.scrollbar_visible = False
        self.scrollbar_hide_timer = None
        self.scrollbar_fade_alpha = 0  # 0-255 透明度
        self.scrollbar_color = '#5a5a5a'  # 默认滚动条颜色
        self.scrollbar_hover_color = '#7a7a7a'  # 悬停时颜色
        self.scrollbar_active_color = '#4a9eff'  # 活动时颜色（蓝色）
        self.scrollbar_is_hovered = False

        # 绑定Canvas事件
        self.image_list_canvas.bind('<Configure>', self._on_canvas_configure)
        self.image_list_canvas.bind('<MouseWheel>', self._on_canvas_mousewheel)
        self.image_list_canvas.bind('<Button-4>', self._on_canvas_mousewheel)
        self.image_list_canvas.bind('<Button-5>', self._on_canvas_mousewheel)
        
        # 鼠标进入/离开事件 - 只在列表区域滚动
        self.image_list_canvas.bind('<Enter>', self._on_list_enter)
        self.image_list_canvas.bind('<Leave>', self._on_list_leave)

        # 绑定滚动条事件
        self.scrollbar_track.bind('<Button-1>', self._on_scrollbar_click)
        self.scrollbar_track.bind('<B1-Motion>', self._on_scrollbar_drag)
        self.scrollbar_track.bind('<Enter>', self._on_scrollbar_enter)
        self.scrollbar_track.bind('<Leave>', self._on_scrollbar_leave)

        # 内部框架用于放置按钮 - 设置宽度为180
        self.image_list_frame = tk.Frame(self.image_list_canvas, bg='#3b3b3b', width=180)
        self.image_list_canvas.create_window((0, 0), window=self.image_list_frame, anchor='nw')

        # 不使用bind_all，改为只在鼠标进入列表时绑定
        self.mouse_in_list = False

        # 滚动条相关变量
        self.scrollbar_dragging = False
        self.scrollbar_last_y = 0

    def _on_canvas_configure(self, event=None):
        """Canvas配置改变时更新滚动区域"""
        self.image_list_frame.update_idletasks()
        self.image_list_canvas.configure(scrollregion=self.image_list_canvas.bbox("all"))
        self._check_scrollbar_needed()

    def _check_scrollbar_needed(self):
        """检查是否需要滚动条"""
        self.image_list_frame.update_idletasks()
        canvas_height = self.image_list_canvas.winfo_height()
        frame_height = self.image_list_frame.winfo_height()
        return frame_height > canvas_height

    def _show_scrollbar(self):
        """显示滚动条（带动画效果）"""
        if not self._check_scrollbar_needed():
            return
            
        # 取消隐藏定时器
        if self.scrollbar_hide_timer:
            self.root.after_cancel(self.scrollbar_hide_timer)
            self.scrollbar_hide_timer = None
        
        if not self.scrollbar_visible:
            self.scrollbar_visible = True
            self.image_list_scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 2))
        
        self._update_scrollbar_thumb()
        
        # 设置隐藏定时器（1.5秒后隐藏）
        self._schedule_hide_scrollbar()
    
    def _schedule_hide_scrollbar(self):
        """安排隐藏滚动条"""
        if self.scrollbar_hide_timer:
            self.root.after_cancel(self.scrollbar_hide_timer)
        
        # 如果正在拖动或悬停，不隐藏
        if not self.scrollbar_dragging and not self.scrollbar_is_hovered:
            self.scrollbar_hide_timer = self.root.after(1500, self._hide_scrollbar)
    
    def _hide_scrollbar(self):
        """隐藏滚动条"""
        if self.scrollbar_visible and not self.scrollbar_dragging and not self.scrollbar_is_hovered:
            self.scrollbar_visible = False
            self.image_list_scrollbar_frame.pack_forget()
    
    def _update_scrollbar_thumb(self):
        """更新滚动条滑块"""
        self.image_list_frame.update_idletasks()

        canvas_height = self.image_list_canvas.winfo_height()
        frame_height = self.image_list_frame.winfo_height()

        if frame_height <= canvas_height:
            if self.scrollbar_thumb_id:
                self.scrollbar_track.delete(self.scrollbar_thumb_id)
                self.scrollbar_thumb_id = None
            return

        # 计算滚动条滑块的大小和位置
        track_height = self.scrollbar_track.winfo_height() - 10  # 减去上下padding
        thumb_height = max(30, int(track_height * canvas_height / frame_height))

        # 获取当前滚动位置
        scroll_fraction = self.image_list_canvas.yview()[0]
        max_scroll = 1.0 - (canvas_height / frame_height)
        if max_scroll > 0:
            thumb_y = int(scroll_fraction / max_scroll * (track_height - thumb_height)) + 5
        else:
            thumb_y = 5

        # 删除旧的滑块
        if self.scrollbar_thumb_id:
            self.scrollbar_track.delete(self.scrollbar_thumb_id)

        # 选择颜色
        if self.scrollbar_dragging:
            color = self.scrollbar_active_color
        elif self.scrollbar_is_hovered:
            color = self.scrollbar_hover_color
        else:
            color = self.scrollbar_color

        # 绘制新的滑块 - 圆角矩形效果
        x1, y1 = 1, thumb_y
        x2, y2 = 5, thumb_y + thumb_height
        radius = 2
        
        # 使用圆角矩形
        self.scrollbar_thumb_id = self.scrollbar_track.create_rectangle(
            x1, y1, x2, y2,
            fill=color, outline=color, width=0
        )

    def _on_canvas_mousewheel(self, event):
        """Canvas鼠标滚轮事件"""
        if not self._check_scrollbar_needed():
            return
            
        if event.num == 4 or event.delta > 0:
            self.image_list_canvas.yview_scroll(-3, "units")
        elif event.num == 5 or event.delta < 0:
            self.image_list_canvas.yview_scroll(3, "units")

        self._show_scrollbar()
    
    def _on_list_enter(self, event):
        """鼠标进入列表区域"""
        self.mouse_in_list = True
        # 绑定全局滚轮事件
        self.root.bind("<MouseWheel>", self._on_global_mousewheel)
        self.root.bind("<Button-4>", self._on_global_mousewheel)
        self.root.bind("<Button-5>", self._on_global_mousewheel)
    
    def _on_list_leave(self, event):
        """鼠标离开列表区域"""
        self.mouse_in_list = False
        # 解绑全局滚轮事件
        self.root.unbind("<MouseWheel>")
        self.root.unbind("<Button-4>")
        self.root.unbind("<Button-5>")
        
        # 安排隐藏滚动条
        self._schedule_hide_scrollbar()
    
    def _on_global_mousewheel(self, event):
        """全局鼠标滚轮事件（只在鼠标在列表区域时生效）"""
        if self.mouse_in_list:
            self._on_canvas_mousewheel(event)

    def _on_scrollbar_click(self, event):
        """滚动条点击事件"""
        canvas_height = self.image_list_canvas.winfo_height()
        frame_height = self.image_list_frame.winfo_height()

        if frame_height > canvas_height:
            self.scrollbar_dragging = True
            # 计算点击位置对应的滚动位置
            track_height = self.scrollbar_track.winfo_height() - 10
            scroll_fraction = (event.y - 5) / track_height
            scroll_fraction = max(0, min(1, scroll_fraction))
            
            max_scroll = 1.0 - (canvas_height / frame_height)
            self.image_list_canvas.yview_moveto(scroll_fraction * max_scroll)
            self._update_scrollbar_thumb()

    def _on_scrollbar_drag(self, event):
        """滚动条拖动事件"""
        canvas_height = self.image_list_canvas.winfo_height()
        frame_height = self.image_list_frame.winfo_height()

        if frame_height > canvas_height:
            self.scrollbar_dragging = True
            track_height = self.scrollbar_track.winfo_height() - 10
            scroll_fraction = (event.y - 5) / track_height
            scroll_fraction = max(0, min(1, scroll_fraction))
            
            max_scroll = 1.0 - (canvas_height / frame_height)
            self.image_list_canvas.yview_moveto(scroll_fraction * max_scroll)
            self._update_scrollbar_thumb()
    
    def _on_scrollbar_enter(self, event):
        """鼠标进入滚动条"""
        self.scrollbar_is_hovered = True
        self._update_scrollbar_thumb()
        # 取消隐藏定时器
        if self.scrollbar_hide_timer:
            self.root.after_cancel(self.scrollbar_hide_timer)
            self.scrollbar_hide_timer = None
    
    def _on_scrollbar_leave(self, event):
        """鼠标离开滚动条"""
        self.scrollbar_is_hovered = False
        self.scrollbar_dragging = False
        self._update_scrollbar_thumb()
        self._schedule_hide_scrollbar()

    def update_image_list_display(self):
        """更新图片列表显示"""
        # 如果数量相同，只更新样式而不重建（避免闪烁）
        if len(self.image_buttons) == len(self.project.images):
            self._update_list_styles_only()
            return
        
        # 数量不同时才重建
        for btn in self.image_buttons:
            btn.destroy()
        self.image_buttons = []

        # 创建新按钮
        for idx, img_data in enumerate(self.project.images):
            self.create_image_list_item(idx, img_data)

        # 更新滚动区域
        self.image_list_frame.update_idletasks()
        self.image_list_canvas.configure(scrollregion=self.image_list_canvas.bbox("all"))
        
        # 检查是否需要滚动条（不自动显示，只在滚动时显示）
        self._check_scrollbar_needed()
    
    def _update_list_styles_only(self):
        """只更新列表项的样式，不重建（避免闪烁）"""
        for idx, item_frame in enumerate(self.image_buttons):
            if idx >= len(self.project.images):
                break
            
            img_data = self.project.images[idx]
            is_selected = (idx == self.project.current_index)
            
            # 更新框架背景色
            bg_color = '#4a4a4a' if is_selected else '#3b3b3b'
            item_frame.config(bg=bg_color)
            
            # 更新子组件样式
            for child in item_frame.winfo_children():
                if isinstance(child, tk.Label):
                    # 文件名标签
                    if hasattr(child, 'cget') and 'text' in child.keys():
                        text = child.cget('text')
                        if '○' in text or '✓' in text:
                            child.config(
                                bg=bg_color,
                                fg='#00ff00' if is_selected else '#cccccc'
                            )
                elif isinstance(child, tk.Frame):
                    # 缩略图容器 - 不改变颜色
                    pass
    
    def _bind_scroll_to_widget(self, widget):
        """为组件绑定滚轮事件（递归绑定所有子组件）"""
        widget.bind('<MouseWheel>', self._on_canvas_mousewheel)
        widget.bind('<Button-4>', self._on_canvas_mousewheel)
        widget.bind('<Button-5>', self._on_canvas_mousewheel)
        
        # 递归绑定子组件
        for child in widget.winfo_children():
            self._bind_scroll_to_widget(child)

    def create_image_list_item(self, idx, img_data):
        """创建图片列表项"""
        # 创建框架 - 固定宽度和高度
        item_frame = tk.Frame(self.image_list_frame, bg='#4a4a4a' if idx == self.project.current_index else '#3b3b3b', 
                              width=175, height=140)
        item_frame.pack(fill=tk.X, padx=5, pady=3)
        item_frame.pack_propagate(False)  # 固定大小

        # 统一缩略图容器尺寸 - 固定为 160x90
        THUMB_WIDTH = 160
        THUMB_HEIGHT = 90
        
        thumb_label = None
        thumb_container = None
        try:
            img = cv2.imdecode(np.fromfile(img_data.path, dtype=np.uint8), cv2.IMREAD_COLOR)
            if img is not None:
                h, w = img.shape[:2]
                
                # 使用适应模式（fit）- 保持图片完整，不裁剪
                scale = min(THUMB_WIDTH / w, THUMB_HEIGHT / h)
                new_w, new_h = int(w * scale), int(h * scale)
                img = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_AREA)

                # 转换为PIL
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                pil_img = Image.fromarray(img_rgb)
                
                # 创建固定尺寸的背景图（深灰色填充）
                background = Image.new('RGB', (THUMB_WIDTH, THUMB_HEIGHT), (45, 45, 45))
                # 将图片居中粘贴到背景上
                paste_x = (THUMB_WIDTH - new_w) // 2
                paste_y = (THUMB_HEIGHT - new_h) // 2
                background.paste(pil_img, (paste_x, paste_y))
                
                photo = ImageTk.PhotoImage(background)

                # 显示缩略图 - 添加边框效果
                thumb_container = tk.Frame(item_frame, bg='#555555', padx=1, pady=1)
                thumb_container.pack(pady=(5, 3))
                
                thumb_label = tk.Label(thumb_container, image=photo, bg='#2d2d2d')
                thumb_label.image = photo  # 保持引用
                thumb_label.pack()
        except Exception as e:
            # 如果加载失败，显示占位符
            print(f"加载缩略图失败 {img_data.name}: {e}")
            thumb_container = tk.Frame(item_frame, bg='#555555', width=THUMB_WIDTH+2, height=THUMB_HEIGHT+2)
            thumb_container.pack(pady=(5, 3))
            thumb_container.pack_propagate(False)
            thumb_label = tk.Label(thumb_container, text="[无法加载]", bg='#2d2d2d', fg='#666666')
            thumb_label.pack(fill=tk.BOTH, expand=True)

        # 文件名和状态
        status_text = "✓" if img_data.is_analyzed else "○"
        saved_text = "💾" if img_data.is_saved else ""
        
        # 截断文件名
        display_name = img_data.name
        if len(display_name) > 18:
            display_name = display_name[:15] + "..."
        
        name_label = tk.Label(
            item_frame,
            text=f"{status_text} {display_name} {saved_text}",
            bg='#4a4a4a' if idx == self.project.current_index else '#3b3b3b',
            fg='#00ff00' if idx == self.project.current_index else '#cccccc',
            font=('Microsoft YaHei', 8),
            anchor='center'
        )
        name_label.pack(pady=(0, 3), fill=tk.X)

        # 收集所有需要绑定悬停效果的组件
        hover_widgets = [item_frame, name_label]
        if thumb_label:
            hover_widgets.append(thumb_label)
        if thumb_container:
            hover_widgets.append(thumb_container)

        # 悬停效果 - 为所有组件统一处理
        def on_enter(e):
            if idx != self.project.current_index:
                item_frame.config(bg='#505050')
                name_label.config(bg='#505050')
        
        def on_leave(e):
            # 检查鼠标是否真的离开了整个item区域
            x, y = item_frame.winfo_pointerxy()
            widget_x = item_frame.winfo_rootx()
            widget_y = item_frame.winfo_rooty()
            widget_w = item_frame.winfo_width()
            widget_h = item_frame.winfo_height()
            
            # 如果鼠标还在item_frame范围内，不改变颜色
            if widget_x <= x <= widget_x + widget_w and widget_y <= y <= widget_y + widget_h:
                return
                
            if idx != self.project.current_index:
                item_frame.config(bg='#3b3b3b')
                name_label.config(bg='#3b3b3b')

        # 点击事件
        def on_click(e=None):
            self.switch_to_image(idx)

        # 为所有组件绑定事件
        for widget in hover_widgets:
            widget.bind("<Button-1>", on_click)
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
        
        # 为所有子组件绑定滚轮事件（解决图片上无法滚动的问题）
        self._bind_scroll_to_widget(item_frame)

        self.image_buttons.append(item_frame)

    def switch_to_image(self, idx):
        """切换到指定图片"""
        # 退出手动交点模式
        self.exit_intersection_mode(restore_image=False)

        if not self.project.set_current_index(idx):
            return

        img_data = self.project.get_current_image()
        if img_data is None:
            return

        try:
            # 加载图片
            self.analyzer.load_image(img_data.path)
            self.current_image = self.analyzer.image.copy()
            self.image_path = img_data.path

            # 从 ImageData 恢复分析数据（使用保存的副本，避免引用问题）
            if img_data.is_analyzed and img_data.center is not None:
                self.analyzer.center = img_data.center
                self.analyzer.radius = img_data.radius
                self.analyzer.inner_radius = img_data.inner_radius
                self.analyzer.results = img_data.results.copy() if img_data.results else []
                if img_data.px_per_mm is not None:
                    self.analyzer.px_per_mm = img_data.px_per_mm

            # 更新 img_data.analyzer 引用
            img_data.analyzer = self.analyzer

            # 恢复结果显示
            self.update_results(self.analyzer.results)

            # 如果已分析，重新绘制标注
            if img_data.is_analyzed and self.analyzer.center is not None:
                result_img = self.analyzer.draw_result()
                if result_img is not None:
                    self.current_image = result_img

            # 重置缩放和偏移
            self.zoom_scale = 1.0
            self.image_offset_x = 0
            self.image_offset_y = 0

            # 清除 PIL 缓存，确保显示正确的图片
            self._pil_cache = None

            # 显示图片
            self.display_current_image()

            # 更新列表显示
            self.update_image_list_display()

            # 更新统计信息
            self.update_project_stats()

            self.status_var.set(f"已切换到: {img_data.name}")
        except Exception as e:
            messagebox.showerror("错误", f"加载图片失败: {e}")

    def update_project_stats(self):
        """更新项目统计信息"""
        total = len(self.project.images)
        analyzed = self.project.get_analyzed_count()
        saved = self.project.get_saved_count()
        stats_text = f"总数: {total} | 已分析: {analyzed} | 已保存: {saved}"
        self.project_stats_label.config(text=stats_text)

    def setup_ui(self):
        """设置界面"""
        # 顶部工具栏
        toolbar = ttk.Frame(self.root)
        toolbar.pack(fill=tk.X, padx=5, pady=5)

        # 按钮样式
        style = ttk.Style()
        style.configure('TButton', font=('Microsoft YaHei', 10))

        # 上传图片按钮
        ttk.Button(toolbar, text="上传图片", command=self.load_image).pack(side=tk.LEFT, padx=5)

        # 批量导入按钮
        ttk.Button(toolbar, text="批量导入", command=self.batch_import_images).pack(side=tk.LEFT, padx=5)

        # DICOM校准按钮
        ttk.Button(toolbar, text="从DICOM校准", command=self.calibrate_dicom).pack(side=tk.LEFT, padx=5)

        # 标尺校准按钮
        ttk.Button(toolbar, text="标尺校准", command=self.start_ruler_calibration).pack(side=tk.LEFT, padx=5)

        # px/mm 输入框（可显示默认值，也可手动修改）
        ttk.Label(toolbar, text="px/mm:").pack(side=tk.LEFT, padx=(5, 0))
        self.pxmm_entry = ttk.Entry(toolbar, textvariable=self.px_per_mm, width=6)
        self.pxmm_entry.pack(side=tk.LEFT)
        self.pxmm_entry.bind("<FocusIn>", lambda e: self.exit_intersection_mode())
        self.pxmm_source_label = ttk.Label(toolbar, text="(默认)", foreground='gray',
                                           font=('Microsoft YaHei', 9))
        self.pxmm_source_label.pack(side=tk.LEFT, padx=(2, 5))

        # 参数输入
        ttk.Label(toolbar, text="内缩r-").pack(side=tk.LEFT, padx=(5, 0))
        self.inner_offset_entry = ttk.Entry(toolbar, textvariable=self.inner_offset, width=5)
        self.inner_offset_entry.pack(side=tk.LEFT)
        self.inner_offset_entry.bind("<FocusIn>", lambda e: self.exit_intersection_mode())
        ttk.Label(toolbar, text="mm").pack(side=tk.LEFT)

        ttk.Label(toolbar, text="ROI半径:").pack(side=tk.LEFT, padx=(5, 0))
        self.roi_radius_entry = ttk.Entry(toolbar, textvariable=self.roi_radius, width=5)
        self.roi_radius_entry.pack(side=tk.LEFT)
        self.roi_radius_entry.bind("<FocusIn>", lambda e: self.exit_intersection_mode())
        ttk.Label(toolbar, text="mm").pack(side=tk.LEFT)

        # 分析按钮
        ttk.Button(toolbar, text="自动分析", command=self.auto_analyze).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="手动交点", command=self.manual_select_intersection).pack(side=tk.LEFT, padx=5)

        # 保存和导出
        ttk.Button(toolbar, text="保存图片", command=self.save_image).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="批量保存", command=self.batch_save_images).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="导出Excel", command=self.export_excel_report).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="清除标注", command=self.clear_annotations).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="重置", command=self.reset).pack(side=tk.LEFT, padx=5)

        # 主内容区域
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 左侧：图片列表面板
        self.setup_image_list_panel(main_frame)

        # 中间：图像显示区
        self.canvas_frame = ttk.Frame(main_frame)
        self.canvas_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(self.canvas_frame, bg='#1e1e1e', highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        self.canvas.bind("<Configure>", self.on_canvas_resize)
        self.canvas.bind("<MouseWheel>", self.on_mouse_wheel)  # Windows滚轮
        self.canvas.bind("<Button-4>", self.on_mouse_wheel)    # Linux滚轮上
        self.canvas.bind("<Button-5>", self.on_mouse_wheel)    # Linux滚轮下

        # 右侧：结果面板
        self.result_frame = ttk.Frame(main_frame, width=280)
        self.result_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10)
        self.result_frame.pack_propagate(False)

        # 结果标题
        ttk.Label(self.result_frame, text="分析结果", font=('Microsoft YaHei', 14, 'bold')).pack(pady=10)

        # 项目统计
        self.project_stats_label = ttk.Label(self.result_frame, text="", font=('Microsoft YaHei', 9), foreground='#ffff00')
        self.project_stats_label.pack(pady=5)

        # 灰度比值显示
        ttk.Label(self.result_frame, text="灰度比值 (硬化带/坏死区)",
                  font=('Microsoft YaHei', 10)).pack(pady=5)

        self.ratio_labels = []
        for i in range(2):  # 只有2个交点
            label = ttk.Label(self.result_frame, text="", font=('Microsoft YaHei', 12), foreground='cyan')
            label.pack(pady=2)
            self.ratio_labels.append(label)

        ttk.Separator(self.result_frame, orient='horizontal').pack(fill=tk.X, pady=10)

        # 平均灰度比
        ttk.Label(self.result_frame, text="平均灰度比", font=('Microsoft YaHei', 10)).pack(pady=5)
        self.avg_ratio_label = ttk.Label(self.result_frame, text="--",
                                          font=('Microsoft YaHei', 24, 'bold'), foreground='#00ff00')
        self.avg_ratio_label.pack(pady=10)

        ttk.Separator(self.result_frame, orient='horizontal').pack(fill=tk.X, pady=10)

        # 图例
        ttk.Label(self.result_frame, text="图例:", font=('Microsoft YaHei', 11, 'bold')).pack(pady=5, anchor='w')

        legends = [
            ("● 绿色大圆 = 股骨头", "#00ff00"),
            ("● 红色十字 = 圆心", "#ff0000"),
            ("● 青色圆 = r-6mm圆", "#00ffff"),
            ("● 橙色点 = 硬化带交点", "#ffa500"),
            ("● 绿色小圆 = 硬化带ROI", "#00ff00"),
            ("● 红色小圆 = 坏死区ROI", "#ff0000"),
        ]
        
        for text, color in legends:
            ttk.Label(self.result_frame, text=text, foreground=color, 
                      font=('Microsoft YaHei', 9)).pack(anchor='w', padx=10)
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, 
                                     font=('Microsoft YaHei', 9), foreground='#00ff00')
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=5)
    
    def load_image(self):
        """加载图片（同时加入图片列表）"""
        # 退出手动交点模式
        self.exit_intersection_mode()

        path = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[("图片文件", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"),
                       ("所有文件", "*.*")]
        )
        if path:
            try:
                # 检查是否已在列表中
                existing_paths = [img.path for img in self.project.images]
                if path not in existing_paths:
                    # 添加到项目
                    self.project.add_images([path])
                    
                    # 创建输出文件夹（如果还没有）
                    if not self.project.output_dir:
                        self.project.create_output_dir()
                    
                    # 设置为当前图片
                    new_index = len(self.project.images) - 1
                    self.project.current_index = new_index
                    
                    # 更新列表显示
                    self.update_image_list_display()
                    self.update_project_stats()
                else:
                    # 如果已存在，切换到该图片
                    new_index = existing_paths.index(path)
                    self.project.current_index = new_index
                
                # 加载并显示图片
                self.image_path = path
                self.analyzer = FemoralHeadAnalyzer(px_per_mm=self.px_per_mm.get())
                self.analyzer.load_image(path)
                self.current_image = self.analyzer.image.copy()
                self.zoom_scale = 1.0
                self.image_offset_x = 0
                self.image_offset_y = 0
                self._pil_cache = None
                self.display_current_image()
                
                # 关联analyzer到ImageData
                img_data = self.project.get_current_image()
                if img_data:
                    img_data.analyzer = self.analyzer
                
                self.update_image_list_display()
                self.status_var.set(f"已加载: {os.path.basename(path)}")
                self.clear_results()
            except Exception as e:
                messagebox.showerror("错误", f"无法加载图片: {e}")

    def batch_import_images(self):
        """批量导入图片"""
        # 退出手动交点模式
        self.exit_intersection_mode()

        paths = filedialog.askopenfilenames(
            title="选择多张图片",
            filetypes=[("图片文件", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff"),
                       ("所有文件", "*.*")]
        )
        if paths:
            try:
                # 创建输出文件夹
                if not self.project.output_dir:
                    self.project.create_output_dir()

                # 添加图片到项目
                self.project.add_images(list(paths))

                # 设置当前图片为第一张
                if self.project.images:
                    self.project.current_index = 0
                    self.switch_to_image(0)

                # 更新列表显示
                self.update_image_list_display()
                self.update_project_stats()

                self.status_var.set(f"已导入 {len(paths)} 张图片")
            except Exception as e:
                messagebox.showerror("错误", f"导入失败: {e}")

    def batch_save_images(self):
        """批量保存所有已分析的图片"""
        # 退出手动交点模式
        self.exit_intersection_mode()

        if not self.project.images:
            messagebox.showwarning("警告", "没有图片可保存")
            return
        
        # 统计已分析的图片数量
        analyzed_images = [img for img in self.project.images if img.is_analyzed]
        
        if not analyzed_images:
            messagebox.showwarning("警告", "没有已分析的图片可保存\n请先对图片进行分析")
            return

        # 创建输出文件夹
        if not self.project.output_dir:
            self.project.create_output_dir()

        if not self.project.output_dir:
            messagebox.showerror("错误", "无法创建输出文件夹")
            return

        saved_count = 0

        for img_data in self.project.images:
            # 只保存已分析的图片
            if img_data.is_analyzed and img_data.center is not None:
                try:
                    # 重新加载原始图片
                    original_img = cv2.imdecode(np.fromfile(img_data.path, dtype=np.uint8), cv2.IMREAD_COLOR)
                    if original_img is None:
                        print(f"无法加载图片 {img_data.name}")
                        continue

                    # 使用保存的分析数据绘制结果
                    vis = original_img.copy()
                    cx, cy = img_data.center

                    # 股骨头（绿色）
                    cv2.circle(vis, img_data.center, img_data.radius, (0, 255, 0), 2)

                    # 圆心（红色十字）
                    cv2.line(vis, (cx-10, cy), (cx+10, cy), (0, 0, 255), 2)
                    cv2.line(vis, (cx, cy-10), (cx, cy+10), (0, 0, 255), 2)

                    # r-6mm内圆（青色）
                    if img_data.inner_radius:
                        cv2.circle(vis, img_data.center, img_data.inner_radius, (255, 255, 0), 2)

                    # 交点和ROI
                    for r in img_data.results:
                        inter = r['intersection']
                        # 交点（橙色）
                        cv2.circle(vis, (inter['x'], inter['y']), 4, (0, 165, 255), -1)
                        # 硬化带ROI（绿色，线宽1）
                        cv2.circle(vis, r['sclerotic'], r['roi_radius'], (0, 255, 0), 1)
                        # 坏死区ROI（红色，线宽1）
                        cv2.circle(vis, r['necrosis'], r['roi_radius'], (0, 0, 255), 1)

                    # 保存图片
                    filename = os.path.splitext(img_data.name)[0] + "_result.png"
                    filepath = os.path.join(self.project.output_dir, filename)
                    # 使用 imencode + tofile 支持中文路径
                    _, img_encoded = cv2.imencode('.png', vis)
                    img_encoded.tofile(filepath)
                    img_data.is_saved = True
                    saved_count += 1
                except Exception as e:
                    print(f"保存 {img_data.name} 失败: {e}")

        self.update_image_list_display()
        self.update_project_stats()
        
        # 构建提示消息
        not_analyzed = len(self.project.images) - len(analyzed_images)
        msg = f"已保存 {saved_count} 张图片到:\n{self.project.output_dir}"
        if not_analyzed > 0:
            msg += f"\n\n{not_analyzed} 张图片未分析，未保存"
        
        messagebox.showinfo("批量保存完成", msg)
        self.status_var.set(f"已保存 {saved_count} 张图片")

    def export_excel_report(self):
        """导出Excel分析报告"""
        # 退出手动交点模式
        self.exit_intersection_mode()

        if not self.project.images:
            messagebox.showwarning("警告", "没有数据可导出")
            return

        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "需要安装openpyxl库\n请运行: pip install openpyxl")
            return

        # 创建输出文件夹
        if not self.project.output_dir:
            self.project.create_output_dir()

        if not self.project.output_dir:
            messagebox.showerror("错误", "无法创建输出文件夹")
            return

        # 生成Excel文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(self.project.output_dir, f"analysis_report_{timestamp}.xlsx")

        success, message = self.project.export_to_excel(excel_path)

        if success:
            messagebox.showinfo("成功", message)
            self.status_var.set(f"已导出Excel报告: {excel_path}")
        else:
            messagebox.showerror("错误", message)

    def display_current_image(self, fast=False):
        """显示当前图像
        fast=True: 使用快速缩放（滚轮缩放时）
        fast=False: 使用高质量缩放（最终渲染）
        """
        if self.current_image is None:
            self.canvas.delete("all")
            self.close_btn_coords = None
            return

        # 缓存颜色转换后的PIL图像（避免每次都转换）
        cache_id = id(self.current_image)
        if self._pil_cache is None or self._cache_id != cache_id:
            img_rgb = cv2.cvtColor(self.current_image, cv2.COLOR_BGR2RGB)
            self._pil_cache = Image.fromarray(img_rgb)
            self._cache_id = cache_id

        img_pil = self._pil_cache

        # 缩放适应画布
        canvas_w = self.canvas.winfo_width()
        canvas_h = self.canvas.winfo_height()

        if canvas_w > 1 and canvas_h > 1:
            img_w, img_h = img_pil.size
            base_scale = min(canvas_w / img_w, canvas_h / img_h, 1.0)
            scale = base_scale * self.zoom_scale
            new_w = int(img_w * scale)
            new_h = int(img_h * scale)
            
            # 限制最大渲染尺寸以提高性能（最大4000像素）
            max_dim = 4000
            if new_w > max_dim or new_h > max_dim:
                limit_scale = max_dim / max(new_w, new_h)
                new_w = int(new_w * limit_scale)
                new_h = int(new_h * limit_scale)
            
            if new_w > 0 and new_h > 0:
                # 快速模式用 NEAREST
                # 高质量模式：小图用 LANCZOS，大图用 BILINEAR（更快）
                if fast:
                    resample = Image.Resampling.NEAREST
                elif new_w * new_h > 2000000:  # 超过200万像素用BILINEAR
                    resample = Image.Resampling.BILINEAR
                else:
                    resample = Image.Resampling.LANCZOS
                img_pil = img_pil.resize((new_w, new_h), resample)

        self.display_image = ImageTk.PhotoImage(img_pil)

        self.canvas.delete("all")

        # 计算图片位置（包含拖动偏移）
        img_x = canvas_w // 2 + self.image_offset_x
        img_y = canvas_h // 2 + self.image_offset_y

        self.canvas.create_image(img_x, img_y, anchor=tk.CENTER, image=self.display_image, tags="image")

        # 绘制关闭按钮（右上角X图标）
        self.draw_close_button()
    
    def on_canvas_resize(self, event):
        """画布大小改变时重绘"""
        if self.current_image is not None:
            self.display_current_image()

    def draw_close_button(self):
        """绘制关闭按钮（右上角X图标）"""
        if self.current_image is None:
            self.close_btn_coords = None
            return

        # 按钮位置和大小
        btn_size = 24
        padding = 10
        x1 = self.canvas.winfo_width() - padding - btn_size
        y1 = padding
        x2 = x1 + btn_size
        y2 = y1 + btn_size

        # 保存按钮坐标用于点击检测
        self.close_btn_coords = (x1, y1, x2, y2)

        # 绘制圆形背景
        self.canvas.create_oval(x1, y1, x2, y2, fill='#ff4444', outline='#cc0000', width=2, tags="close_btn")

        # 绘制X
        margin = 7
        self.canvas.create_line(x1 + margin, y1 + margin, x2 - margin, y2 - margin,
                               fill='white', width=2, tags="close_btn")
        self.canvas.create_line(x1 + margin, y2 - margin, x2 - margin, y1 + margin,
                               fill='white', width=2, tags="close_btn")

    def on_mouse_wheel(self, event):
        """鼠标滚轮缩放（更平滑）"""
        if self.current_image is None:
            return

        # 退出手动交点模式
        self.exit_intersection_mode()
        # 获取滚轮方向，使用更小的步长使缩放更平滑
        if event.num == 4 or (hasattr(event, 'delta') and event.delta > 0):
            # 放大：乘以1.1（10%）
            self.zoom_scale *= 1.1
        elif event.num == 5 or (hasattr(event, 'delta') and event.delta < 0):
            # 缩小：乘以0.9（10%）
            self.zoom_scale *= 0.9

        # 限制缩放范围 10%-300%
        self.zoom_scale = max(0.1, min(3.0, self.zoom_scale))

        # 快速渲染（低质量，不卡顿）
        self.display_current_image(fast=True)
        self.status_var.set(f"缩放: {self.zoom_scale:.0%}")

        # 取消之前的高质量渲染定时器
        if self._zoom_timer is not None:
            self.root.after_cancel(self._zoom_timer)

        # 停止滚动 200ms 后再高质量渲染
        self._zoom_timer = self.root.after(200, self._zoom_finalize)

    def _zoom_finalize(self):
        """缩放结束后高质量渲染"""
        self._zoom_timer = None
        self.display_current_image(fast=False)

    def close_image(self):
        """关闭当前图片"""
        self.current_image = None
        self.display_image = None
        self.image_path = None
        self.analyzer = FemoralHeadAnalyzer()
        self.zoom_scale = 1.0
        self.image_offset_x = 0  # 重置偏移
        self.image_offset_y = 0
        self.close_btn_coords = None
        self.manual_mode = False
        self.intersection_mode = False
        self.intersection_points = []
        self.necrosis_direction_point = None
        self.ruler_mode = False
        self.ruler_points = []
        self.click_points = []
        self._pil_cache = None
        self._cache_id = None

        # 清空画布和结果
        self.canvas.delete("all")
        self.clear_results()
        self.pxmm_source_label.config(text="(默认)", foreground='gray')
        self.px_per_mm.set(5.0)
        self.status_var.set("图片已关闭")
    
    def on_canvas_click(self, event):
        """画布按下事件"""
        # 检查是否点击了关闭按钮
        if self.close_btn_coords is not None:
            x1, y1, x2, y2 = self.close_btn_coords
            if x1 <= event.x <= x2 and y1 <= event.y <= y2:
                self.close_image()
                return

        # 检查是否点击在图片范围内
        self.click_on_image = False
        if self.current_image is not None:
            canvas_w = self.canvas.winfo_width()
            canvas_h = self.canvas.winfo_height()
            img_h, img_w = self.current_image.shape[:2]
            
            base_scale = min(canvas_w / img_w, canvas_h / img_h, 1.0)
            scale = base_scale * self.zoom_scale
            disp_w = int(img_w * scale)
            disp_h = int(img_h * scale)
            
            # 图片在画布上的边界
            img_left = (canvas_w - disp_w) // 2 + self.image_offset_x
            img_top = (canvas_h - disp_h) // 2 + self.image_offset_y
            img_right = img_left + disp_w
            img_bottom = img_top + disp_h
            
            # 检查点击是否在图片范围内
            if img_left <= event.x <= img_right and img_top <= event.y <= img_bottom:
                self.click_on_image = True

        # 记录拖动起始位置
        self.drag_start_x = event.x
        self.drag_start_y = event.y
        self.is_dragging = False

    def on_canvas_drag(self, event):
        """画布拖动事件"""
        if self.current_image is None:
            return
        
        # 只有点击在图片上才能拖动
        if not self.click_on_image:
            return

        # 计算移动距离
        dx = event.x - self.drag_start_x
        dy = event.y - self.drag_start_y

        # 如果移动超过3像素，认为是拖动（更灵敏）
        if not self.is_dragging and (abs(dx) > 3 or abs(dy) > 3):
            self.is_dragging = True
            self.canvas.config(cursor="fleur")  # 切换为移动光标
            self.canvas.delete("close_btn")     # 隐藏关闭按钮
            # 退出手动交点模式
            self.exit_intersection_mode()

        if self.is_dragging:
            # 更新图片偏移
            self.image_offset_x += dx
            self.image_offset_y += dy
            self.drag_start_x = event.x
            self.drag_start_y = event.y
            # 只移动图片，不移动关闭按钮
            self.canvas.move("image", dx, dy)

    def on_canvas_release(self, event):
        """画布释放事件"""
        # 如果是拖动，恢复光标并重绘
        if self.is_dragging:
            self.is_dragging = False
            self.canvas.config(cursor="")  # 恢复默认光标
            self.draw_close_button()       # 重绘关闭按钮（固定位置）
            return

        # 以下是点击逻辑
        if self.analyzer.gray is None:
            return

        # 计算实际图像坐标
        canvas_w = self.canvas.winfo_width()
        canvas_h = self.canvas.winfo_height()
        img_h, img_w = self.analyzer.gray.shape

        base_scale = min(canvas_w / img_w, canvas_h / img_h, 1.0)
        scale = base_scale * self.zoom_scale  # 考虑用户缩放
        disp_w = int(img_w * scale)
        disp_h = int(img_h * scale)

        # 考虑图片偏移
        offset_x = (canvas_w - disp_w) // 2 + self.image_offset_x
        offset_y = (canvas_h - disp_h) // 2 + self.image_offset_y

        img_x = int((event.x - offset_x) / scale)
        img_y = int((event.y - offset_y) / scale)

        if not (0 <= img_x < img_w and 0 <= img_y < img_h):
            return

        # 标尺校准模式
        if self.ruler_mode:
            self.ruler_points.append((img_x, img_y))

            if len(self.ruler_points) == 1:
                # 绘制第一个红点
                self.draw_ruler_marks()
                self.status_var.set("已选择标尺起点，请点击标尺终点...")
            elif len(self.ruler_points) >= 2:
                # 绘制第二个红点和连线
                self.draw_ruler_marks()

                # 计算像素距离
                x1, y1 = self.ruler_points[0]
                x2, y2 = self.ruler_points[1]
                pixel_distance = np.sqrt((x2-x1)**2 + (y2-y1)**2)

                # 弹窗输入实际距离
                self.ask_ruler_distance(pixel_distance)

                self.ruler_mode = False
                self.ruler_points = []
            return

        # 手动选择模式
        if self.manual_mode:
            self.click_points.append((img_x, img_y))

            if len(self.click_points) == 1:
                self.status_var.set("已选择圆心，请点击股骨头边缘确定半径...")
            elif len(self.click_points) >= 2:
                # 计算半径
                cx, cy = self.click_points[0]
                ex, ey = self.click_points[1]
                radius = int(np.sqrt((ex-cx)**2 + (ey-cy)**2))

                self.analyzer.set_femoral_head((cx, cy), radius)

                # 使用当前px/mm值进行分析（不自动计算）
                self.do_analysis()

                self.manual_mode = False
                self.click_points = []
            return

        # 手动交点模式（2个交点 + 1个坏死区方向）
        if self.intersection_mode:
            if len(self.intersection_points) < 2:
                # 收集交点
                self.intersection_points.append((img_x, img_y))
                self.draw_intersection_marks()

                if len(self.intersection_points) == 1:
                    self.status_var.set("手动交点：请点击第2个硬化带交点...")
                elif len(self.intersection_points) == 2:
                    self.status_var.set("手动交点：请点击坏死区域（确定方向）...")
            else:
                # 第3次点击 - 坏死区方向
                self.necrosis_direction_point = (img_x, img_y)
                self.draw_intersection_marks()  # 绘制坏死区标记

                # 执行分析，返回是否成功
                success = self.do_manual_intersection_analysis()

                # 只有分析成功才退出手动交点模式
                if success:
                    self.intersection_mode = False
                    self.intersection_points = []
                    self.necrosis_direction_point = None
            return

    def draw_intersection_marks(self):
        """绘制手动选择的交点和坏死区标记"""
        if self.intersection_backup_image is None:
            return

        # 从备份图像复制
        img = self.intersection_backup_image.copy()

        # 绘制交点（橙色，半径4像素）
        for (x, y) in self.intersection_points:
            cv2.circle(img, (x, y), 4, (0, 165, 255), -1, cv2.LINE_AA)

        # 绘制坏死区方向点（红色，半径4像素）
        if self.necrosis_direction_point is not None:
            nx, ny = self.necrosis_direction_point
            cv2.circle(img, (nx, ny), 4, (0, 0, 255), -1, cv2.LINE_AA)
            # 从两个交点中心画一条线到坏死区点
            if len(self.intersection_points) == 2:
                mid_x = (self.intersection_points[0][0] + self.intersection_points[1][0]) // 2
                mid_y = (self.intersection_points[0][1] + self.intersection_points[1][1]) // 2
                cv2.line(img, (mid_x, mid_y), (nx, ny), (0, 0, 255), 1, cv2.LINE_AA)

        # 更新显示
        self.current_image = img
        self._pil_cache = None  # 清除缓存
        self.display_current_image()

    def do_manual_intersection_analysis(self):
        """根据手动选择的交点和坏死区方向执行分析
        
        逻辑：
        1. 两个交点在内圆（r-内缩值）上
        2. 根据当前px/mm和内缩参数，计算内圆半径
        3. 根据两个交点位置，反推圆心（在交点连线的垂直平分线上）
        4. 圆心在坏死区的反方向
        
        返回：True 成功，False 失败
        """
        if len(self.intersection_points) < 2 or self.necrosis_direction_point is None:
            return False

        self.status_var.set("正在计算圆心位置...")
        self.root.update()

        # 获取当前参数（使用用户设置的值，不自动重新计算）
        px_mm = self.px_per_mm.get()
        inner_offset_mm = self.inner_offset.get()  # 内缩值，默认6mm
        
        # 假设平均股骨头半径22mm，内圆半径 = (22 - inner_offset_mm) mm
        avg_femoral_radius_mm = 22.0
        inner_radius_mm = avg_femoral_radius_mm - inner_offset_mm  # 16mm（如果内缩6mm）
        inner_radius_px = inner_radius_mm * px_mm  # 内圆半径（像素）
        outer_radius_px = avg_femoral_radius_mm * px_mm  # 外圆半径（像素）

        # 两个交点
        p1 = np.array(self.intersection_points[0])
        p2 = np.array(self.intersection_points[1])
        
        # 计算两个交点的中点（弦的中点）
        mid_point = (p1 + p2) / 2
        
        # 计算弦长
        chord_length = np.linalg.norm(p2 - p1)
        half_chord = chord_length / 2
        
        # 检查弦长是否合理（不能超过内圆直径）
        if half_chord > inner_radius_px:
            messagebox.showwarning("警告", 
                f"两个交点距离过大，无法构成有效的圆\n"
                f"当前交点距离: {chord_length:.0f}像素\n"
                f"内圆直径: {inner_radius_px*2:.0f}像素\n"
                f"请检查px/mm设置或重新选择交点")
            
            # 清除标注，恢复原始图像
            if self.intersection_backup_image is not None:
                self.current_image = self.intersection_backup_image.copy()
                self._pil_cache = None
                self.display_current_image()
            
            # 重置交点数据
            self.intersection_points = []
            self.necrosis_direction_point = None
            self.status_var.set("手动交点：请点击第1个硬化带交点...")
            return False  # 返回失败
        
        # 计算圆心到弦的距离：d = sqrt(r² - (chord/2)²)
        center_to_chord_dist = np.sqrt(inner_radius_px**2 - half_chord**2)
        
        # 计算弦的垂直方向（单位向量）
        chord_vec = p2 - p1
        # 垂直向量（逆时针旋转90度）
        perp_vec = np.array([-chord_vec[1], chord_vec[0]])
        perp_unit = perp_vec / np.linalg.norm(perp_vec)
        
        # 坏死区方向
        necrosis_point = np.array(self.necrosis_direction_point)
        necrosis_dir = necrosis_point - mid_point
        
        # 判断圆心应该在哪一侧（在坏死区的反方向）
        # 计算坏死区方向与垂直向量的点积，确定圆心方向
        dot_product = np.dot(necrosis_dir, perp_unit)
        
        # 圆心在坏死区的反方向
        if dot_product > 0:
            center = mid_point - perp_unit * center_to_chord_dist
        else:
            center = mid_point + perp_unit * center_to_chord_dist
        
        # 设置圆心和半径
        cx, cy = int(center[0]), int(center[1])
        self.analyzer.center = (cx, cy)
        self.analyzer.radius = int(outer_radius_px)
        self.analyzer.inner_radius = int(inner_radius_px)

        # 计算ROI参数
        roi_radius_px = max(3, int(self.roi_radius.get() * px_mm))
        # 偏移量 = 橙色点半径(4) + ROI半径，让测量圆边缘刚好挨着橙色点边缘
        orange_radius = 4
        offset_px = orange_radius + roi_radius_px

        # 构建结果 - 每个交点使用从圆心指向该交点的径向方向
        self.analyzer.results = []
        for (ix, iy) in self.intersection_points:
            # 计算从圆心指向该交点的径向方向（单位向量）
            radial_dx = ix - cx
            radial_dy = iy - cy
            radial_len = np.sqrt(radial_dx**2 + radial_dy**2)
            if radial_len > 0:
                radial_ux = radial_dx / radial_len
                radial_uy = radial_dy / radial_len
            else:
                radial_ux, radial_uy = 0, -1

            # 硬化带ROI（沿径向向内偏移，靠近圆心）
            sclerotic_x = int(ix - radial_ux * offset_px)
            sclerotic_y = int(iy - radial_uy * offset_px)

            # 坏死区ROI（沿径向向外偏移，远离圆心，与硬化带对称）
            necrosis_x = int(ix + radial_ux * offset_px)
            necrosis_y = int(iy + radial_uy * offset_px)

            g_necrosis = self.analyzer._calc_roi_gray(necrosis_x, necrosis_y, roi_radius_px)
            g_sclerotic = self.analyzer._calc_roi_gray(sclerotic_x, sclerotic_y, roi_radius_px)

            ratio = g_sclerotic / g_necrosis if g_necrosis > 0 else 0

            self.analyzer.results.append({
                'intersection': {'x': ix, 'y': iy},
                'necrosis': (necrosis_x, necrosis_y),
                'sclerotic': (sclerotic_x, sclerotic_y),
                'roi_radius': roi_radius_px,
                'g_necrosis': g_necrosis,
                'g_sclerotic': g_sclerotic,
                'ratio': ratio
            })

        # 绘制结果
        self.current_image = self.analyzer.draw_result()
        self._pil_cache = None  # 清除缓存
        self.display_current_image()

        # 更新结果显示
        self.update_results(self.analyzer.results)

        # 保存结果到当前ImageData（与自动分析保持一致）
        if self.project.current_index >= 0:
            img_data = self.project.get_current_image()
            if img_data:
                img_data.results = self.analyzer.results.copy() if self.analyzer.results else []
                img_data.is_analyzed = True
                img_data.analyzer = self.analyzer
                img_data.px_per_mm_source = self.pxmm_source_label.cget("text")
                # 保存分析数据的副本（避免引用问题）
                img_data.center = self.analyzer.center
                img_data.radius = self.analyzer.radius
                img_data.inner_radius = self.analyzer.inner_radius
                img_data.px_per_mm = self.analyzer.px_per_mm
                self.update_image_list_display()
                self.update_project_stats()
        
        self.status_var.set(f"手动交点分析完成 | px/mm={px_mm:.2f} | 内缩r-{inner_offset_mm}mm | ROI={self.roi_radius.get()}mm")
        return True  # 返回成功

    def start_ruler_calibration(self):
        """开始标尺校准"""
        if self.analyzer.gray is None:
            messagebox.showwarning("警告", "请先加载图片")
            return

        # 退出手动交点模式（恢复原图）
        self.exit_intersection_mode()

        self.ruler_mode = True
        self.ruler_points = []
        self.manual_mode = False  # 退出手动圆心模式
        # 保存原始图像，用于校准完成后恢复
        self.ruler_backup_image = self.current_image.copy() if self.current_image is not None else None
        self.status_var.set("标尺校准：请点击标尺起点...")

    def draw_ruler_marks(self):
        """在图像上绘制标尺校准的红点和红线"""
        if self.ruler_backup_image is None or len(self.ruler_points) == 0:
            return

        # 从备份图像复制，避免累加绘制
        img = self.ruler_backup_image.copy()

        # 如果有两个点，绘制红线（先画线再画点，点在上层）
        if len(self.ruler_points) >= 2:
            x1, y1 = self.ruler_points[0]
            x2, y2 = self.ruler_points[1]
            cv2.line(img, (x1, y1), (x2, y2), (0, 0, 255), 1, cv2.LINE_AA)  # 抗锯齿红线，线宽1

        # 绘制红点（半径4像素）
        for (x, y) in self.ruler_points:
            cv2.circle(img, (x, y), 4, (0, 0, 255), -1, cv2.LINE_AA)  # 红色实心圆，抗锯齿

        # 更新显示
        self.current_image = img
        self.display_current_image()

    def ask_ruler_distance(self, pixel_distance):
        """弹窗输入标尺实际距离"""
        dialog = tk.Toplevel(self.root)
        dialog.title("输入标尺距离")

        # 设置弹窗大小
        dialog_width = 320
        dialog_height = 180

        # 获取屏幕尺寸，计算居中靠上位置
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - dialog_width) // 2
        y = screen_height // 4  # 靠上位置（屏幕高度的1/4处）

        dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"测量的像素距离: {pixel_distance:.1f} 像素",
                  font=('Microsoft YaHei', 10)).pack(pady=10)
        ttk.Label(dialog, text="请输入这两点之间的实际距离 (mm):").pack(pady=5)

        entry = ttk.Entry(dialog, width=10)
        entry.insert(0, "100")  # 默认100mm
        entry.pack(pady=5)
        entry.select_range(0, tk.END)
        entry.focus()

        def confirm():
            try:
                mm_distance = float(entry.get())
                if mm_distance > 0:
                    # 计算 px/mm
                    px_mm = pixel_distance / mm_distance
                    self.px_per_mm.set(round(px_mm, 2))
                    self.analyzer.px_per_mm = px_mm
                    self.pxmm_source_label.config(text="(标尺)", foreground='#ffff00')
                    self.status_var.set(f"标尺校准完成: {px_mm:.2f} px/mm")
                    # 恢复原始图像（移除红点红线）
                    restore_image()
                    dialog.destroy()
                else:
                    messagebox.showerror("错误", "请输入正数")
            except ValueError:
                messagebox.showerror("错误", "请输入有效数字")

        def restore_image():
            """恢复原始图像"""
            if self.ruler_backup_image is not None:
                self.current_image = self.ruler_backup_image.copy()
                self.display_current_image()
                self.ruler_backup_image = None

        def on_cancel():
            """取消时恢复原始图像"""
            restore_image()
            self.status_var.set("标尺校准已取消")
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)  # 处理关闭按钮

        ttk.Button(dialog, text="确定", command=confirm).pack(pady=10)
        entry.bind("<Return>", lambda e: confirm())

    def calibrate_dicom(self):
        """从DICOM校准"""
        # 退出手动交点模式
        self.exit_intersection_mode()

        try:
            import pydicom
            path = filedialog.askopenfilename(
                title="选择DICOM文件",
                filetypes=[("DICOM文件", "*.dcm"), ("所有文件", "*.*")]
            )
            if path:
                ds = pydicom.dcmread(path)
                if hasattr(ds, 'PixelSpacing'):
                    spacing = float(ds.PixelSpacing[0])
                    px_mm = 1.0 / spacing
                    self.px_per_mm.set(round(px_mm, 2))
                    self.analyzer.px_per_mm = px_mm
                    self.pxmm_source_label.config(text="(DICOM)", foreground='#00ff00')
                    self.status_var.set(f"已从DICOM校准: {px_mm:.2f} px/mm")
                else:
                    messagebox.showwarning("警告", "DICOM文件中没有PixelSpacing信息")
        except ImportError:
            messagebox.showerror("错误", "需要安装pydicom库\npip install pydicom")
        except Exception as e:
            messagebox.showerror("错误", f"读取DICOM失败: {e}")

    def auto_calibrate_pxmm(self):
        """根据平均股骨头直径(44mm)自动校准px/mm"""
        # 退出手动交点模式
        self.exit_intersection_mode()

        if self.analyzer.radius is None:
            return

        # 成年人平均股骨头直径 44mm，半径 22mm
        avg_radius_mm = 22.0
        detected_radius_px = self.analyzer.radius

        # 计算 px/mm = 像素半径 / 毫米半径
        px_mm = detected_radius_px / avg_radius_mm

        # 更新值
        self.px_per_mm.set(round(px_mm, 2))
        self.analyzer.px_per_mm = px_mm
        self.pxmm_source_label.config(text="(自动)", foreground='cyan')
        self.status_var.set(f"已按平均股骨头直径(44mm)校准: {px_mm:.2f} px/mm")

    def auto_analyze(self):
        """自动分析"""
        if self.analyzer.gray is None:
            messagebox.showwarning("警告", "请先加载图片")
            return

        # 退出手动交点模式
        self.exit_intersection_mode()

        self.status_var.set("正在检测股骨头...")
        self.root.update()

        # 尝试自动检测
        result = self.analyzer.detect_femoral_head()

        if result is None:
            messagebox.showwarning("警告", "自动检测失败，请使用手动选择")
            return

        # 使用当前px/mm值进行分析（不自动计算）
        self.do_analysis()

    def manual_select_center(self):
        """手动选择圆心模式"""
        if self.analyzer.gray is None:
            messagebox.showwarning("警告", "请先加载图片")
            return

        self.manual_mode = True
        self.intersection_mode = False
        self.ruler_mode = False
        self.click_points = []
        self.status_var.set("手动圆心：请点击股骨头圆心...")

    def exit_intersection_mode(self, restore_image=True):
        """退出手动交点模式

        Args:
            restore_image: 是否恢复备份图像（默认True，当用户未完成交点选择时恢复原图）
        """
        if not self.intersection_mode:
            return

        # 如果需要恢复图像且有备份图像，则恢复
        if restore_image and self.intersection_backup_image is not None:
            self.current_image = self.intersection_backup_image.copy()
            self._pil_cache = None
            self.display_current_image()
            self.status_var.set("已退出手动交点模式")

        # 清除状态
        self.intersection_mode = False
        self.intersection_points = []
        self.necrosis_direction_point = None
        self.intersection_backup_image = None

    def manual_select_intersection(self):
        """手动选择交点模式（2个交点 + 1个坏死区方向）"""
        if self.analyzer.gray is None:
            messagebox.showwarning("警告", "请先加载图片")
            return

        self.intersection_mode = True
        self.manual_mode = False
        self.ruler_mode = False
        self.intersection_points = []
        self.necrosis_direction_point = None  # 坏死区方向点
        # 保存当前图像用于绘制交点
        self.intersection_backup_image = self.analyzer.image.copy()
        self.current_image = self.intersection_backup_image.copy()
        self._pil_cache = None
        self.display_current_image()
        self.status_var.set("手动交点：请点击第1个硬化带交点...")

    def do_analysis(self):
        """执行分析"""
        self.status_var.set("正在分析...")
        self.root.update()

        # 更新参数
        self.analyzer.px_per_mm = self.px_per_mm.get()

        # 执行分析
        results = self.analyzer.analyze(
            inner_offset_mm=self.inner_offset.get(),
            roi_offset_mm=2.0,
            roi_radius_mm=self.roi_radius.get()
        )

        # 绘制结果
        self.current_image = self.analyzer.draw_result()
        self.display_current_image()

        # 更新结果显示
        self.update_results(results)

        # 保存结果到当前ImageData
        if self.project.current_index >= 0:
            img_data = self.project.get_current_image()
            if img_data:
                img_data.results = results.copy() if results else []
                img_data.is_analyzed = True
                img_data.analyzer = self.analyzer
                img_data.px_per_mm_source = self.pxmm_source_label.cget("text")
                # 保存分析数据的副本（避免引用问题）
                img_data.center = self.analyzer.center
                img_data.radius = self.analyzer.radius
                img_data.inner_radius = self.analyzer.inner_radius
                img_data.px_per_mm = self.analyzer.px_per_mm
                self.update_image_list_display()
                self.update_project_stats()

        self.status_var.set(f"分析完成 | px/mm={self.px_per_mm.get():.2f} | 内缩r-{self.inner_offset.get()}mm | ROI={self.roi_radius.get()}mm")
    
    def update_results(self, results):
        """更新结果显示"""
        # 清空之前的结果
        for label in self.ratio_labels:
            label.config(text="")
        
        # 显示每个交点的结果（只有2个交点）
        for i, r in enumerate(results[:2]):
            text = f"交点{i+1}: {r['ratio']:.4f}"
            self.ratio_labels[i].config(text=text)
        
        # 平均值
        if results:
            avg = np.mean([r['ratio'] for r in results])
            self.avg_ratio_label.config(text=f"{avg:.4f}")
        else:
            self.avg_ratio_label.config(text="--")
    
    def clear_results(self):
        """清空结果"""
        for label in self.ratio_labels:
            label.config(text="")
        self.avg_ratio_label.config(text="--")
    
    def save_image(self):
        """保存结果图片（只能保存已分析的图片）"""
        # 退出手动交点模式
        self.exit_intersection_mode()

        if self.current_image is None:
            messagebox.showwarning("警告", "没有可保存的图片")
            return
        
        # 检查是否已分析
        img_data = self.project.get_current_image()
        if img_data and not img_data.is_analyzed:
            messagebox.showwarning("警告", "当前图片尚未分析，请先进行分析后再保存")
            return
        
        # 如果没有在项目中但有分析结果，也允许保存
        if not img_data and (self.analyzer.results is None or len(self.analyzer.results) == 0):
            messagebox.showwarning("警告", "当前图片尚未分析，请先进行分析后再保存")
            return

        # 如果有项目，保存到项目输出文件夹
        if self.project.images and self.project.current_index >= 0:
            if not self.project.output_dir:
                self.project.create_output_dir()

            if self.project.output_dir:
                if img_data:
                    filename = os.path.splitext(img_data.name)[0] + "_result.png"
                    path = os.path.join(self.project.output_dir, filename)
                    cv2.imwrite(path, self.current_image)
                    img_data.is_saved = True
                    self.update_image_list_display()
                    self.update_project_stats()
                    # 弹窗提示保存位置
                    messagebox.showinfo("保存成功", f"图片已保存到:\n{path}")
                    self.status_var.set(f"已保存: {path}")
                    return

        # 否则使用文件对话框
        path = filedialog.asksaveasfilename(
            title="保存图片",
            defaultextension=".png",
            filetypes=[("PNG图片", "*.png"), ("JPEG图片", "*.jpg"), ("所有文件", "*.*")]
        )
        if path:
            cv2.imwrite(path, self.current_image)
            messagebox.showinfo("保存成功", f"图片已保存到:\n{path}")
            self.status_var.set(f"已保存: {path}")
    
    def clear_annotations(self):
        """清除图片上的标注（保留原图）"""
        if self.analyzer.image is None:
            messagebox.showwarning("警告", "没有已加载的图片")
            return

        # 如果在手动交点模式下，只清除当前标注，重新开始
        if self.intersection_mode:
            # 清除已点击的点
            self.intersection_points = []
            self.necrosis_direction_point = None
            # 恢复备份图像
            if self.intersection_backup_image is not None:
                self.current_image = self.intersection_backup_image.copy()
                self._pil_cache = None
                self.display_current_image()
            self.status_var.set("手动交点：请点击第1个硬化带交点...")
            return

        # 恢复原始图像
        self.current_image = self.analyzer.image.copy()
        self._pil_cache = None
        self.display_current_image()

        # 清除分析结果
        self.analyzer.center = None
        self.analyzer.radius = None
        self.analyzer.inner_radius = None
        self.analyzer.results = []

        # 重置所有手动模式
        self.manual_mode = False
        self.click_points = []
        self.intersection_mode = False
        self.intersection_points = []
        self.necrosis_direction_point = None
        self.intersection_backup_image = None
        self.ruler_mode = False
        self.ruler_points = []
        self.ruler_backup_image = None

        # 重置当前图片的分析状态
        img_data = self.project.get_current_image()
        if img_data:
            img_data.is_analyzed = False
            img_data.is_saved = False
            img_data.results = []
            self.update_image_list_display()
            self.update_project_stats()

        # 清空结果显示
        self.clear_results()
        self.status_var.set("已清除标注")

    def reset(self):
        """重置（包括清空图片列表）"""
        self.analyzer = FemoralHeadAnalyzer(px_per_mm=self.px_per_mm.get())
        self.current_image = None
        self.image_path = None
        self.manual_mode = False
        self.intersection_mode = False
        self.intersection_points = []
        self.necrosis_direction_point = None
        self.ruler_mode = False
        self.ruler_points = []
        self.click_points = []
        self._pil_cache = None
        self._cache_id = None
        self.zoom_scale = 1.0
        self.image_offset_x = 0
        self.image_offset_y = 0
        self.close_btn_coords = None
        
        # 清空项目和图片列表
        self.project = ImageProject()
        
        # 清空图片列表显示
        for btn in self.image_buttons:
            btn.destroy()
        self.image_buttons = []
        
        # 更新滚动区域
        self.image_list_frame.update_idletasks()
        self.image_list_canvas.configure(scrollregion=self.image_list_canvas.bbox("all"))
        
        # 隐藏滚动条
        if self.scrollbar_visible:
            self.scrollbar_visible = False
            self.image_list_scrollbar_frame.pack_forget()
        
        self.canvas.delete("all")
        self.clear_results()
        self.update_project_stats()
        self.pxmm_source_label.config(text="(默认)", foreground='gray')
        self.px_per_mm.set(5.0)
        self.status_var.set("已重置")


def main():
    root = tk.Tk()
    
    # 设置主题
    style = ttk.Style()
    style.theme_use('clam')
    
    # 深色主题配置
    style.configure('.', background='#2b2b2b', foreground='white')
    style.configure('TFrame', background='#2b2b2b')
    style.configure('TLabel', background='#2b2b2b', foreground='white')
    style.configure('TButton', background='#404040', foreground='white')
    style.map('TButton', background=[('active', '#505050')])
    style.configure('TEntry', fieldbackground='white', foreground='black')
    
    app = FemoralHeadApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()